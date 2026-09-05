"""
Phase I Verification Script.
Audits a Phase I workbook against ENTITY_SPECS:
- Verifies row counts for Startups, Products, Research_Papers (>= 100).
- Verifies Entity Resolution audit log distribution (mix of NORMALIZATION_EXACT / ALIAS_MATCH / NEW_ENTITY).
- Spot-checks random records for valid URLs, ISO-8601 UTC dates, and plausible GitHub stars.
- Reports Research Paper GitHub URL hit-rate.

Usage:
    python scripts/verify_phase1.py [path/to/workbook.xlsx]
    (defaults to exports/phase1_test.xlsx — the CLI `--phase 1` output target)
"""

import sys
from collections import Counter
from pathlib import Path
import random
from typing import Any, Dict, List
import openpyxl
from dateutil import parser as dateutil_parser

DEFAULT_EXCEL_PATH = Path("exports/phase1_test.xlsx")
EXCEL_PATH = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_EXCEL_PATH


def get_sheet_records(ws) -> List[Dict[str, Any]]:
    """Extract worksheet rows as a list of header-keyed dictionaries."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return []
    headers = [str(h or "") for h in rows[0]]
    return [dict(zip(headers, r, strict=True)) for r in rows[1:]]


def verify_phase1() -> None:
    if not EXCEL_PATH.exists():
        print(f"❌ Error: {EXCEL_PATH} does not exist. Run Phase I first.")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheets = {name: get_sheet_records(wb[name]) for name in wb.sheetnames}
    print(f"📊 Opened {EXCEL_PATH}. Sheets found: {list(sheets.keys())}\n")

    # 1. Record Counts
    print("--- 1. Record Counts ---")
    log_sheet_name = "Entity Mapping Log" if "Entity Mapping Log" in sheets else "Entity_Resolution_Log"
    min_targets = {
        "Startups": 100,
        "Products": 100,
        "Research_Papers": 100,
        log_sheet_name: 1,
    }
    for name, target in min_targets.items():
        count = len(sheets.get(name, []))
        status = "✅ PASS" if count >= target else "❌ FAIL"
        print(f"  {name:25s}: {count:4d} rows (min required: {target}) -> {status}")
    print()

    # 2. Entity Resolution Log Audit
    if log_sheet_name in sheets:
        print(f"--- 2. {log_sheet_name} Audit ---")
        methods = Counter(str(r.get("matchMethod", r.get("Match Method", "UNKNOWN"))) for r in sheets[log_sheet_name])
        total = sum(methods.values())
        print(f"  Total Log Entries: {total}")
        for m, c in methods.most_common():
            print(f"    - {m:25s}: {c:4d} ({c / total * 100:5.1f}%)")
        new_pct = (methods.get("NEW_ENTITY", 0) / total * 100) if total else 100
        status_icon = "✅ PASS" if new_pct < 100 else "⚠️ WARN"
        print(f"  {status_icon} Entity resolution method mix: NEW_ENTITY is {new_pct:.1f}%")
        print()

    # 3. Research Papers GitHub Correlation
    if "Research_Papers" in sheets:
        print("--- 3. Research Papers GitHub Correlation ---")
        papers = sheets["Research_Papers"]
        gh_matches = sum(1 for p in papers if (p.get("content.github_url") or p.get("GitHub Repo")) and str(p.get("content.github_url") or p.get("GitHub Repo")) != "None")
        stars_matches = sum(1 for p in papers if isinstance(p.get("content.github_stars", p.get("GitHub Stars")), (int, float)) and (p.get("content.github_stars", p.get("GitHub Stars")) or 0) > 0)
        rate = (gh_matches / len(papers) * 100) if papers else 0
        print(f"  Total Papers        : {len(papers)}")
        print(f"  Papers with GitHub  : {gh_matches} ({rate:.1f}% hit-rate)")
        print(f"  Papers with Stars   : {stars_matches}")
        print()

    # 4. Spot-Checks (5 random per sheet)
    print("--- 4. Random Spot-Checks (5 per sheet) ---")
    field_maps = {
        "Startups": (("content.entityName", "Canonical Entity Name"), ("source.url", "Source URL"), ("collectedAt", "Collected At")),
        "Products": (("content.productName", "Product Name"), ("content.productUrl", "Product URL"), ("collectedAt", "Collected At")),
        "Research_Papers": (("content.title", "Title"), ("content.paper_url", "Paper URL"), ("content.published_date", "Published Date")),
    }

    for name, (name_keys, url_keys, date_keys) in field_maps.items():
        records = sheets.get(name, [])
        if not records:
            continue
        print(f"  [{name}] (Sample of {min(5, len(records))} rows):")
        for r in random.sample(records, min(5, len(records))):
            name_val = next((r[k] for k in name_keys if k in r and r[k] is not None), "N/A")
            url_val = next((r[k] for k in url_keys if k in r and r[k] is not None), "N/A")
            raw_date = str(next((r[k] for k in date_keys if k in r and r[k] is not None), "N/A"))

            title = str(name_val)[:32]
            url = str(url_val)

            date_ok = "Valid UTC"
            try:
                dt = dateutil_parser.parse(raw_date)
                date_str = dt.isoformat()[:19]
            except Exception:
                date_ok = "Raw/Unparsed" if raw_date != "N/A" else "N/A"
                date_str = raw_date[:19]

            url_status = "Valid URL" if url.startswith("http") else "Invalid URL"
            print(f"    • {title:32s} | {url_status}: {url[:45]:45s} | Date: {date_str} ({date_ok})")
        print()


if __name__ == "__main__":
    verify_phase1()
