"""
Tests for the CLI orchestration layer: end-to-end run_pipeline (crawlers -> exports
-> summary -> run report), the click entry point, and the reporting/monitor helpers.
Follows AAA + offline conventions; all network is replaced at the crawler seam so the
test exercises the real orchestration, exporters, and report machinery.
"""

import asyncio
import json
from pathlib import Path
from types import SimpleNamespace

import csv

import pytest
import openpyxl
import respx
import httpx
from click.testing import CliRunner
from freezegun import freeze_time

from src import cli
from src.schemas.entities import (
    JobContent,
    JobRecord,
    NewsContent,
    NewsRecord,
    PricingModelEnum,
    ProductContent,
    ProductRecord,
    ResearchPaperContent,
    ResearchPaperRecord,
    RoleFamilyEnum,
    SourceMetadata,
    StartupContent,
    StartupRecord,
)


def _make_records() -> dict:
    """One valid record per vertical, built through the real pydantic schemas."""
    src = lambda name: SourceMetadata(name=name, url=f"https://{name.lower()}.example")
    return {
        "startups": [
            StartupRecord(
                source=src("Y Combinator"),
                content=StartupContent(entityName="Acme AI"),
            )
        ],
        "products": [
            ProductRecord(
                source=src("Product Hunt"),
                content=ProductContent(
                    startupName="Acme AI",
                    productName="Acme Copilot",
                    pricingModel=PricingModelEnum.FREEMIUM,
                ),
            )
        ],
        "papers": [
            ResearchPaperRecord(
                content=ResearchPaperContent(
                    title="Scaling AI Agents",
                    authors=["Ada Lovelace"],
                    paper_url="https://arxiv.org/abs/2401.0001",
                    github_url="https://github.com/acme/agents",
                    github_stars=42,
                    published_date="2026-09-04T10:00:00Z",
                ),
            )
        ],
        "news": [
            NewsRecord(
                source=src("TechCrunch AI"),
                content=NewsContent(
                    title="Acme AI raises a new round",
                    published_date="2026-09-04T10:00:00Z",
                    summary="Funding news for Acme AI.",
                    full_text="Acme AI announced a new funding round today with strong investor demand.",
                ),
            )
        ],
        "jobs": [
            JobRecord(
                source=src("RemoteOK AI"),
                content=JobContent(
                    company="Acme AI",
                    title="AI Engineer",
                    date="2026-09-04T10:00:00Z",
                    is_remote=True,
                    role_family=RoleFamilyEnum.ENGINEERING,
                ),
            )
        ],
    }


from src.crawlers.base import TargetedCrawler


class _FakeCrawler(TargetedCrawler):
    """Async-context-manager crawler stand-in returning a fixed record list."""

    def __init__(self, records, **kwargs):
        super().__init__(**kwargs)
        self.records = records
        self.collected = list(records)

    async def __aenter__(self):
        return self

    async def __aexit__(self, *exc):
        return False

    async def crawl(self):
        self.recover_from_wal()
        return list(self.records)


def _patch_crawler_classes(monkeypatch, records: dict):
    """Route cli's five crawler classes to fixed-record fakes (per-vertical)."""
    for name in ("papers", "startups", "products", "news", "jobs"):
        cls_name = {
            "papers": "ResearchPapersCrawler",
            "startups": "StartupsCrawler",
            "products": "ProductsCrawler",
            "news": "NewsCrawler",
            "jobs": "JobsCrawler",
        }[name]
        fake_cls = type(cls_name, (_FakeCrawler,), {})
        monkeypatch.setattr(
            cli, cls_name,
            lambda records=records[name], fake_cls=fake_cls, **kwargs: fake_cls(records, **kwargs),
        )


@pytest.mark.asyncio
async def test_run_pipeline_end_to_end_full_target(tmp_path, monkeypatch):
    # Arrange - a complete run with every vertical returning exactly one record and a
    # target of one: all five crawlers run concurrently, exports write real files,
    # the summary renders, and the run report records 'completed'.
    _patch_crawler_classes(monkeypatch, _make_records())
    monkeypatch.setattr(cli, "entity_resolver", SimpleNamespace(save_cache=lambda: None, audit_log=[]))
    out_xlsx = str(tmp_path / "pipeline.xlsx")

    # Act
    success = await cli.run_pipeline(
        run_phase1=True, run_phase2=True, target_count=1,
        output_xlsx=out_xlsx, progress_interval=0,
    )

    # Assert - completed, workbook written with all six tabs, report persisted
    assert success is True
    wb = openpyxl.load_workbook(out_xlsx)
    assert set(wb.sheetnames) == {
        "Startups", "Products", "Research_Papers", "Jobs_24h", "News_24h", "Entity Mapping Log",
    }
    assert wb["Startups"].max_row == 2  # header + one record
    report = json.loads(open(tmp_path / "run_report.json").read())
    assert report["status"] == "completed"
    assert report["collected"] == {"startups": 1, "products": 1, "papers": 1, "news": 1, "jobs": 1}
    assert report["target_count"] == 1


@pytest.mark.asyncio
async def test_run_pipeline_shortfall_reports_and_exit(tmp_path, monkeypatch):
    # Arrange - target above what any vertical supplies: shortfall must flip the result
    # and the report status, driving the CLI's SystemExit(1).
    _patch_crawler_classes(monkeypatch, _make_records())
    monkeypatch.setattr(cli, "entity_resolver", SimpleNamespace(save_cache=lambda: None, audit_log=[]))

    # Act
    success = await cli.run_pipeline(
        run_phase1=True, run_phase2=False, target_count=10,
        output_xlsx=str(tmp_path / "pipeline.xlsx"), progress_interval=0,
    )

    # Assert
    assert success is False
    report = json.loads(open(tmp_path / "run_report.json").read())
    assert report["status"] == "shortfall"
    assert report["phase2"] is False


def test_cli_main_success_and_shortfall_exit(tmp_path, monkeypatch):
    # Arrange - wire main() to a fake pipeline so the click wiring, config warnings,
    # default output paths, and exit-code contract are all exercised offline.
    captured = {}

    async def fake_run_pipeline(**kwargs):
        captured.update(kwargs)
        return True

    # Spy on the config warnings so presence is asserted without relying on rich
    # output capture inside the click runner.
    original_warn = cli._warn_config
    warn_calls = []

    def spy_warn(*args, **kwargs):
        warn_calls.append(args)
        return original_warn(*args, **kwargs)

    monkeypatch.setattr(cli, "run_pipeline", fake_run_pipeline)
    monkeypatch.setattr(cli, "_warn_config", spy_warn)
    runner = CliRunner()

    # Act 1 - full run reports success and uses the default all-phase output
    result = runner.invoke(cli.main, ["--phase", "all", "--target", "2", "--progress-interval", "0"])
    assert result.exit_code == 0
    assert captured["output_xlsx"] == cli.DEFAULT_OUTPUT_XLSX
    assert captured["target_count"] == 2
    assert captured["run_phase1"] and captured["run_phase2"]
    assert warn_calls and warn_calls[0][0] is True  # config warnings ran for phase 1

    # Act 2 - phase-1 run selects the phase-1 output path
    result = runner.invoke(cli.main, ["--phase", "1", "--target", "2", "--progress-interval", "0"])
    assert result.exit_code == 0
    assert captured["output_xlsx"] == cli.PHASE1_OUTPUT_XLSX
    assert captured["run_phase2"] is False
    assert captured["fresh"] is False

    # Act 3 - --fresh flag sets fresh=True in run_pipeline kwargs
    result = runner.invoke(cli.main, ["--phase", "all", "--target", "1000", "--sheets", "--fresh", "--progress-interval", "0"])
    assert result.exit_code == 0
    assert captured["fresh"] is True
    assert captured["upload_sheets"] is True


def test_cli_main_shortfall_raises_system_exit(tmp_path, monkeypatch):
    async def fake_run_pipeline(**kwargs):
        return False

    monkeypatch.setattr(cli, "run_pipeline", fake_run_pipeline)
    runner = CliRunner()
    result = runner.invoke(cli.main, ["--phase", "all", "--target", "5", "--progress-interval", "0"])
    assert result.exit_code == 1


def test_warn_config_matrices(monkeypatch, capsys):
    # Every warning branch of _warn_config, toggling settings one at a time.
    monkeypatch.setattr(cli.settings, "github_token", "")
    monkeypatch.setattr(cli.settings, "gemini_api_key", "")
    monkeypatch.setattr(cli.settings, "groq_api_key", "")
    monkeypatch.setattr(cli.settings, "custom_llm_api_key", "")
    monkeypatch.setattr(cli.settings, "deepseek_api_key", "")
    monkeypatch.setattr(cli.settings, "google_service_account_path", None)

    cli._warn_config(run_phase1=True, upload_sheets=True, target_count=1500)
    out = capsys.readouterr().out
    assert "GITHUB_TOKEN not set" in out
    assert "No LLM API keys set" in out
    assert "WAL checkpointing enabled" in out
    assert "--sheets requested but GOOGLE_SERVICE_ACCOUNT_PATH is not set" in out

    # Configured variants must stay silent
    monkeypatch.setattr(cli.settings, "github_token", "gh_abc123")
    monkeypatch.setattr(cli.settings, "gemini_api_key", "sk-x")
    monkeypatch.setattr(cli.settings, "google_service_account_path", "credentials/sa.json")
    cli._warn_config(run_phase1=True, upload_sheets=True, target_count=10)
    out = capsys.readouterr().out
    assert "not set" not in out

    # Fresh flag overrides WAL checkpoint message
    cli._warn_config(run_phase1=True, upload_sheets=False, target_count=1500, fresh=True)
    out = capsys.readouterr().out
    assert "--fresh flag set" in out


def test_warn_stale_sources_two_zeros_and_recovery(monkeypatch, capsys):
    history = {
        "news": {
            "TechCrunch AI": {"recent_fresh_counts": [5, 0, 0]},   # stale
            "The Verge AI": {"recent_fresh_counts": [3, 0, 2]},    # recovered
            "Broken entry": {"not": "a dict"},                     # ignored
        }
    }
    monkeypatch.setattr(cli, "load_source_freshness", lambda crawler: history.get(crawler, {}))
    cli._warn_stale_sources()
    out = capsys.readouterr().out
    assert "Stale source: 'TechCrunch AI' (news)" in out
    assert "The Verge AI" not in out  # only a one-run dip


@pytest.mark.asyncio
async def test_crawler_progress_variants(monkeypatch):
    # collected/target crawlers, news stats, jobs live-count, and absent crawler.
    monkeypatch.setattr(cli, "_ACTIVE_CRAWLERS", {})
    assert cli._crawler_progress("missing") == (0, None)

    class WithCollected:
        collected = [1, 2, 3]
        target_count = 10

    monkeypatch.setattr(cli, "_ACTIVE_CRAWLERS", {"startups": WithCollected()})
    assert cli._crawler_progress("startups") == (3, 10)

    class WithStats:
        stats = {"Feed A": {"total": 2}, "Feed B": {"total": 1}}

    monkeypatch.setattr(cli, "_ACTIVE_CRAWLERS", {"news": WithStats()})
    assert cli._crawler_progress("news") == (3, None)

    class WithLiveCount:
        _live_count = 4

    monkeypatch.setattr(cli, "_ACTIVE_CRAWLERS", {"jobs": WithLiveCount()})
    assert cli._crawler_progress("jobs") == (4, None)


@pytest.mark.asyncio
async def test_safe_gather_exception_and_cancellation():
    async def ok():
        return ["a"]

    async def boom():
        raise RuntimeError("crawler exploded")

    results = await cli._safe_gather([ok(), boom()])
    assert results == [["a"], []]

    async def self_cancel():
        raise asyncio.CancelledError()

    with pytest.raises(asyncio.CancelledError):
        await cli._safe_gather([ok(), self_cancel()])


@pytest.mark.asyncio
async def test_progress_monitor_renders_and_cancels(monkeypatch, capsys):
    # A live monitor tick over an active target-bearing crawler, then clean cancel.
    class Active:
        collected = [1]
        target_count = 10

    monkeypatch.setattr(cli, "_ACTIVE_CRAWLERS", {"papers": Active()})
    monitor_task, cache_task = cli._start_background_tasks(0.02, run_phase1=True, run_phase2=False)
    assert monitor_task is not None
    await asyncio.sleep(0.08)
    await cli._cancel_background_tasks(monitor_task, cache_task)
    assert monitor_task.cancelled()
    out = capsys.readouterr().out
    assert "Live Pipeline Progress" in out
    assert "papers" in out


@pytest.mark.asyncio
async def test_progress_monitor_surfaces_papers_enrichment(monkeypatch, capsys):
    # Active papers crawler: 1000/1000 collected, 412/605 stars enriched (68%), 1 token -> ETA ~7 min
    class MockPapersCrawler:
        collected = list(range(1000))
        target_count = 1000
        enriched_count = 412
        enrich_total = 605
        is_enriching = True
        github_tokens = ["ghp_token1"]
        _exhausted_github_tokens = set()

    monkeypatch.setattr(cli, "console", cli.Console(width=120))
    monkeypatch.setattr(cli, "_ACTIVE_CRAWLERS", {"papers": MockPapersCrawler()})
    monitor_task, cache_task = cli._start_background_tasks(0.02, run_phase1=True, run_phase2=False)
    await asyncio.sleep(0.08)
    await cli._cancel_background_tasks(monitor_task, cache_task)
    out = capsys.readouterr().out

    assert "1000/1000 collected" in out
    assert "enriching stars 412/605 (68%)" in out
    assert "ETA ~7 min" in out
    assert " ↳ papers: stars" in out
    # While enrichment is in-flight, papers' overall status must NOT render as complete or ~0 min
    assert "✅ Complete" not in out
    assert "~0 min" not in out


@pytest.mark.asyncio
async def test_progress_monitor_papers_enriching_mid_collection(monkeypatch, capsys):
    # Active papers crawler still collecting: 500/1000 collected, 100/200 stars enriched
    class MockPapersCrawler:
        collected = list(range(500))
        target_count = 1000
        enriched_count = 100
        enrich_total = 200
        is_enriching = True
        github_tokens = ["ghp_token1"]
        _exhausted_github_tokens = set()

    monkeypatch.setattr(cli, "console", cli.Console(width=120))
    monkeypatch.setattr(cli, "_ACTIVE_CRAWLERS", {"papers": MockPapersCrawler()})
    monitor_task, cache_task = cli._start_background_tasks(0.02, run_phase1=True, run_phase2=False)
    await asyncio.sleep(0.08)
    await cli._cancel_background_tasks(monitor_task, cache_task)
    out = capsys.readouterr().out

    assert "500 (50%)" in out
    assert "enriching stars 100/200 (50%)" in out
    assert " ↳ papers: stars" in out


def test_papers_enrichment_details_edge_cases():
    assert cli._papers_enrichment_details(None) is None

    class Inactive:
        is_enriching = False

    assert cli._papers_enrichment_details(Inactive()) is None

    class ZeroTotal:
        is_enriching = True
        enrich_total = 0
        enriched_count = 0

    assert cli._papers_enrichment_details(ZeroTotal()) is None


@pytest.mark.asyncio
async def test_progress_monitor_retains_completed_crawler_counts(monkeypatch, capsys):
    """Assert progress table retains final counts and marks 'Done' after crawlers complete."""
    monkeypatch.setattr(cli, "console", cli.Console(width=120))
    cli._reset_crawler_state()

    cli._mark_crawler_done("startups", 1000, 1000)
    cli._mark_crawler_done("products", 1000, 1000)
    cli._mark_crawler_done("papers", 1000, 1000)
    cli._mark_crawler_done("news", 39, None)
    cli._mark_crawler_done("jobs", 11, None)

    assert len(cli._ACTIVE_CRAWLERS) == 0
    assert cli._is_crawler_done("startups")
    assert cli._is_crawler_done("news")

    monitor_task, cache_task = cli._start_background_tasks(0.02, run_phase1=True, run_phase2=True)
    await asyncio.sleep(0.08)
    await cli._cancel_background_tasks(monitor_task, cache_task)
    out = capsys.readouterr().out

    assert "1000 (100%)" in out
    assert "39" in out
    assert "11" in out
    assert "✅ Done" in out

    lines = [l for l in out.splitlines() if "│" in l]
    for line in lines:
        parts = [p.strip() for p in line.split("│")]
        if len(parts) >= 6 and parts[1] in ("startups", "products", "papers"):
            assert parts[2] == "1000 (100%)"
            assert "— (24h window)" not in line
            assert parts[5] == "✅ Done"
        if len(parts) >= 6 and parts[1] == "news":
            assert parts[2] == "39"
            assert parts[3] == "— (24h window)"
            assert parts[5] == "✅ Done"
        if len(parts) >= 6 and parts[1] == "jobs":
            assert parts[2] == "11"
            assert parts[3] == "— (24h window)"
            assert parts[5] == "✅ Done"


@pytest.mark.asyncio
async def test_progress_monitor_fallback_target_labels_by_vertical(monkeypatch, capsys):
    """Label news/jobs as (24h window) and target verticals as (target) when target is None."""
    monkeypatch.setattr(cli, "console", cli.Console(width=120))
    cli._reset_crawler_state()

    monitor_task, cache_task = cli._start_background_tasks(0.02, run_phase1=True, run_phase2=True)
    await asyncio.sleep(0.08)
    await cli._cancel_background_tasks(monitor_task, cache_task)
    out = capsys.readouterr().out

    lines = out.splitlines()
    for line in lines:
        if any(k in line for k in ("startups", "products", "papers")):
            assert "— (target)" in line
            assert "— (24h window)" not in line
        if any(k in line for k in ("news", "jobs")):
            assert "— (24h window)" in line
            assert "— (target)" not in line


@pytest.mark.asyncio
async def test_crawl_marks_crawler_done_on_completion():
    """_crawl must register completed count and target in _COMPLETED_CRAWLERS upon exit."""
    cli._reset_crawler_state()

    class DummyCrawler:
        def __init__(self, target_count=1000):
            self.target_count = target_count
            self.collected = [1] * 1000

        async def __aenter__(self):
            return self

        async def __aexit__(self, exc_type, exc_val, exc_tb):
            return False

        async def crawl(self):
            return self.collected

    results = await cli._crawl("startups", DummyCrawler, target_count=1000)
    assert len(results) == 1000
    assert "startups" not in cli._ACTIVE_CRAWLERS
    assert cli._is_crawler_done("startups")
    assert cli._crawler_progress("startups") == (1000, 1000)


@pytest.mark.asyncio
async def test_start_background_tasks_disabled_monitor(monkeypatch):
    monkeypatch.setattr(cli, "_ACTIVE_CRAWLERS", {})
    monitor_task, cache_task = cli._start_background_tasks(0.0, True, True)
    assert monitor_task is None
    await cli._cancel_background_tasks(None, cache_task)
    assert cache_task.cancelled()


def test_handle_sheets_upload_unconfigured(monkeypatch, capsys):
    class UnconfiguredExporter:
        def is_configured(self):
            return False

    monkeypatch.setattr(cli, "GoogleSheetsExporter", UnconfiguredExporter)
    cli._handle_sheets_upload(startups=[], products=[], papers=[], jobs=[], news=[], logs=[])
    out = capsys.readouterr().out
    assert "Google Sheets export skipped: Service account credentials not found" in out


def test_handle_sheets_upload_configured_url_and_none(monkeypatch, capsys):
    calls = {}

    class ConfiguredExporter:
        def is_configured(self):
            return True

        def export(self, **kwargs):
            calls["datasets"] = kwargs
            return "https://sheets.google.com/spreadsheets/d/abc"

    monkeypatch.setattr(cli, "GoogleSheetsExporter", ConfiguredExporter)
    cli._handle_sheets_upload(startups=[1], products=[], papers=[], jobs=[], news=[], logs=[])
    assert "Live Google Sheets URL" in capsys.readouterr().out
    assert calls["datasets"]["startups"] == [1]

    # export() returning None: guidance message instead of a URL
    class NoUrlExporter(ConfiguredExporter):
        service_account_email = "sa@project.iam.gserviceaccount.com"

        def export(self, **kwargs):
            return None

    monkeypatch.setattr(cli, "GoogleSheetsExporter", NoUrlExporter)
    cli._handle_sheets_upload(startups=[], products=[], papers=[], jobs=[], news=[], logs=[])
    assert "Google Sheets upload could not complete" in capsys.readouterr().out


# ---------------------------------------------------------------------------
# End-to-end pipeline with a fully faked transport layer: the REAL crawler
# classes run through run_pipeline while respx serves every HTTP endpoint;
# the real exporters write files that are then verified on disk.
# ---------------------------------------------------------------------------


_ARXIV_FEED = """<?xml version="1.0" encoding="UTF-8"?>
<rss version="2.0">
  <channel>
    <title>cs.AI updates</title>
    <item>
      <title>Scaling Inference for Agentic AI Systems. (arXiv:2409.00001)</title>
      <link>https://arxiv.org/abs/2409.00001</link>
      <description>We present new inference techniques for agentic systems.</description>
      <dc:creator xmlns:dc="http://purl.org/dc/elements/1.1/">Ada Lovelace, Alan Turing</dc:creator>
      <pubDate>Thu, 03 Sep 2026 10:00:00 GMT</pubDate>
    </item>
    <item>
      <title>Retrieval Augmented Reasoning at Scale. (arXiv:2409.00002)</title>
      <link>https://arxiv.org/abs/2409.00002</link>
      <description>We study retrieval augmented reasoning across many benchmarks.</description>
      <dc:creator xmlns:dc="http://purl.org/dc/elements/1.1/">Grace Hopper</dc:creator>
      <pubDate>Thu, 03 Sep 2026 11:00:00 GMT</pubDate>
    </item>
  </channel>
</rss>
"""

_PRODUCTS_MD = (
    "# AI Tools\n"
    "- [Acme Copilot](https://copilot.acme.example) - AI coding assistant for engineering teams.\n"
    "- [Acme Search](https://search.acme.example) - Enterprise semantic search platform.\n"
)


def _route_transport() -> None:
    """Register every HTTP endpoint the real crawlers may hit during a small run.
    Phase I crawlers only: arXiv API (papers), one product source document
    (products), and YC page 1 (startups — reached only if the seed list does
    not already fill the quota)."""
    respx.get("https://export.arxiv.org/api/query").mock(return_value=httpx.Response(200, text=_ARXIV_FEED))
    respx.get("https://raw.example/products.md").mock(return_value=httpx.Response(200, text=_PRODUCTS_MD))
    respx.get(url__startswith="https://api.ycombinator.com/v0.1/companies").mock(
        return_value=httpx.Response(
            200, json={"companies": [{"name": "Yc Backup Startup", "website": "https://yc.example"}]}
        )
    )


@pytest.mark.asyncio
@respx.mock
async def test_run_pipeline_end_to_end_real_crawlers_completed(tmp_path, monkeypatch):
    # Arrange - a small completed run: every vertical supplies exactly the target.
    # No WAL (target < 1000), no LLM keys (deterministic tiers), all network faked.
    _route_transport()
    monkeypatch.setattr(cli.settings, "product_sources_json", '[["Test Source", "https://raw.example/products.md"]]')
    out_xlsx = str(tmp_path / "e2e_completed.xlsx")

    # Act - run the REAL pipeline end to end (real crawler classes + exporters)
    success = await cli.run_pipeline(
        run_phase1=True, run_phase2=False, target_count=2,
        output_xlsx=out_xlsx, progress_interval=0,
    )

    # Assert - completed: True, workbook has data rows, CSVs match, report says so
    assert success is True

    wb = openpyxl.load_workbook(out_xlsx)
    assert set(wb.sheetnames) == {
        "Startups", "Products", "Research_Papers", "Jobs_24h", "News_24h", "Entity Mapping Log",
    }
    assert wb["Startups"].max_row == 3     # header + 2 records
    assert wb["Products"].max_row == 3
    assert wb["Research_Papers"].max_row == 3
    assert wb["Jobs_24h"].max_row == 1     # empty in a phase-1-only run
    assert wb["News_24h"].max_row == 1

    for filename in ("startups.csv", "products.csv", "research_papers.csv"):
        with open(tmp_path / filename, newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))
        assert len(rows) == 3, filename  # header + 2 data rows

    report = json.loads(open(tmp_path / "run_report.json").read())
    assert report["status"] == "completed"
    assert report["phase2"] is False
    assert report["collected"] == {"startups": 2, "products": 2, "papers": 2, "news": 0, "jobs": 0}
    assert report["target_count"] == 2


@pytest.mark.asyncio
@respx.mock
async def test_run_pipeline_end_to_end_real_crawlers_shortfall(tmp_path, monkeypatch):
    # Arrange - the products source document 404s: products collects 0, the run
    # completes but must report a shortfall (exit semantics -> False).
    respx.get("https://export.arxiv.org/api/query").mock(return_value=httpx.Response(200, text=_ARXIV_FEED))
    respx.get("https://raw.example/products.md").mock(return_value=httpx.Response(404))
    monkeypatch.setattr(cli.settings, "product_sources_json", '[["Test Source", "https://raw.example/products.md"]]')

    # Act
    success = await cli.run_pipeline(
        run_phase1=True, run_phase2=False, target_count=2,
        output_xlsx=str(tmp_path / "e2e_shortfall.xlsx"), progress_interval=0,
    )

    # Assert - shortfall flips the result and the report status
    assert success is False
    report = json.loads(open(tmp_path / "run_report.json").read())
    assert report["status"] == "shortfall"
    assert report["collected"]["startups"] == 2
    assert report["collected"]["papers"] == 2
    assert report["collected"]["products"] == 0


@freeze_time("2026-09-04T12:00:00Z")
@pytest.mark.asyncio
@respx.mock
async def test_run_pipeline_end_to_end_all_phases_composed(tmp_path, monkeypatch):
    # Arrange - Composed Phase 1 and Phase 2 run to completion with real crawler classes
    # and all external HTTP endpoints intercepted via respx.
    _route_transport()
    monkeypatch.setattr(cli.settings, "product_sources_json", '[["Test Source", "https://raw.example/products.md"]]')

    # Phase 2 News Mocks
    news_rss = """<?xml version="1.0" encoding="UTF-8"?>
    <rss version="2.0">
      <channel>
        <title>AI News</title>
        <item>
          <title>Anthropic Launches Claude 3.5 Sonnet</title>
          <link>https://techcrunch.com/2026/09/04/claude-launch</link>
          <pubDate>Fri, 04 Sep 2026 10:00:00 GMT</pubDate>
          <description>Anthropic announces new benchmark performance.</description>
        </item>
      </channel>
    </rss>"""
    for feed_url in (
        "https://techcrunch.com/category/artificial-intelligence/feed/",
        "https://venturebeat.com/category/ai/feed/",
        "https://arstechnica.com/tag/ai/feed/",
        "https://www.technologyreview.com/topic/artificial-intelligence/feed",
        "https://hnrss.org/frontpage",
    ):
        respx.get(feed_url).mock(return_value=httpx.Response(200, text=news_rss))

    # Phase 2 Job Mocks
    remoteok_payload = [
        {"legal": "Notice"},
        {
            "id": 1,
            "company": "Anthropic",
            "position": "Senior AI Systems Engineer",
            "date": "2026-09-04T10:00:00Z",
            "url": "https://remoteok.com/jobs/1",
            "tags": ["ai", "remote"],
            "location": "Remote",
        },
    ]
    respx.get("https://remoteok.com/api?tag=ai").mock(return_value=httpx.Response(200, json=remoteok_payload))
    respx.get("https://www.arbeitnow.com/api/job-board-api").mock(return_value=httpx.Response(200, json={"data": []}))
    respx.get("https://himalayas.app/jobs/api?q=ai").mock(return_value=httpx.Response(200, json={"jobs": []}))
    respx.get("https://himalayas.app/jobs/rss").mock(return_value=httpx.Response(200, text="<rss></rss>"))
    respx.get(url__startswith="https://weworkremotely.com").mock(return_value=httpx.Response(200, text="<rss></rss>"))
    respx.get("https://hnrss.org/whoishiring/jobs?q=AI").mock(return_value=httpx.Response(200, text="<rss></rss>"))

    out_xlsx = str(tmp_path / "composed_pipeline.xlsx")

    # Act - Run both Phase 1 and Phase 2 composed
    success = await cli.run_pipeline(
        run_phase1=True,
        run_phase2=True,
        target_count=2,
        output_xlsx=out_xlsx,
        progress_interval=0,
    )

    # Assert - Pipeline completes successfully
    assert success is True

    # 1. Open and verify all 6 tabs in workbook
    wb = openpyxl.load_workbook(out_xlsx)
    assert set(wb.sheetnames) == {
        "Startups", "Products", "Research_Papers", "Jobs_24h", "News_24h", "Entity Mapping Log",
    }
    assert wb["Startups"].max_row >= 3
    assert wb["Products"].max_row >= 3
    assert wb["Research_Papers"].max_row >= 3
    assert wb["Jobs_24h"].max_row >= 2
    assert wb["News_24h"].max_row >= 2
    assert wb["Entity Mapping Log"].max_row >= 2

    # 2. Verify all CSV files are created and populated
    for filename in ("startups.csv", "products.csv", "research_papers.csv", "jobs.csv", "news.csv"):
        with open(tmp_path / filename, newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))
        assert len(rows) >= 2, f"CSV {filename} missing data rows"

    # 3. Verify run report reflects complete composed state
    report = json.loads(open(tmp_path / "run_report.json").read())
    assert report["status"] == "completed"
    assert report["phase2"] is True
    assert report["collected"]["startups"] >= 2
    assert report["collected"]["products"] >= 2
    assert report["collected"]["papers"] >= 2
    assert report["collected"]["news"] >= 1
    assert report["collected"]["jobs"] >= 1


@pytest.mark.asyncio
async def test_run_pipeline_fresh_flag_truncates_wal_and_resets_freshness(tmp_path, monkeypatch):
    """With --fresh passed to run_pipeline: WAL files are truncated and per-source freshness history is reset."""
    # Arrange
    wal_dir = tmp_path / "wal"
    wal_dir.mkdir()
    wal_file = wal_dir / "researchpaperscrawler_wal.jsonl"
    wal_file.write_text(json.dumps({"key": "old-paper", "data": {"title": "Old"}}) + "\n", encoding="utf-8")
    assert wal_file.stat().st_size > 0

    state_path = tmp_path / "run_state.json"
    from src.utils.run_state import save_source_freshness, load_source_freshness
    save_source_freshness("news", {"DeadFeed": 0}, state_path=str(state_path))
    save_source_freshness("news", {"DeadFeed": 0}, state_path=str(state_path))
    assert len(load_source_freshness("news", state_path=str(state_path))["DeadFeed"]["recent_fresh_counts"]) == 2

    monkeypatch.setattr(cli.settings, "wal_dir", str(wal_dir))
    monkeypatch.setattr(cli.settings, "run_state_path", str(state_path))
    _patch_crawler_classes(monkeypatch, _make_records())
    monkeypatch.setattr(cli, "entity_resolver", SimpleNamespace(save_cache=lambda: None, audit_log=[]))

    out_xlsx = str(tmp_path / "fresh_run.xlsx")

    # Act
    success = await cli.run_pipeline(
        run_phase1=True, run_phase2=True, target_count=1,
        output_xlsx=out_xlsx, progress_interval=0,
        fresh=True,
    )

    # Assert
    assert success is True
    assert wal_file.read_text(encoding="utf-8") == ""
    freshness = load_source_freshness("news", state_path=str(state_path))
    assert "DeadFeed" not in freshness




def test_run_report_includes_sheets_status_freshness_and_stale_sources(tmp_path, monkeypatch):
    # P5-2: the run report must carry sheets_upload status, per-source fresh counts,
    # and the stale-source list so cron/CI can alert without reading console output.
    from src import cli as cli_mod
    from src.utils import run_state as rs

    monkeypatch.setattr("src.config.settings.run_state_path", str(tmp_path / "run_state.json"))
    rs.save_source_freshness("news", {"HealthyFeed": 4, "DeadFeed": 0})
    rs.save_source_freshness("news", {"HealthyFeed": 3, "DeadFeed": 0})

    path = cli_mod._write_run_report(
        run_id="test-run",
        duration_s=1.0,
        status="completed",
        counts={"startups": 1, "products": 1, "papers": 1, "news": 2, "jobs": 0},
        target_count=1,
        resolution_log_rows=0,
        phase1=True,
        phase2=True,
        sheets_upload="skipped",
        output_dir=str(tmp_path),
    )
    report = json.loads(Path(path).read_text())

    assert report["sheets_upload"] == "skipped"
    assert report["source_freshness"]["news"]["HealthyFeed"] == 3
    stale = {(s["crawler"], s["source"]) for s in report["stale_sources"]}
    assert ("news", "DeadFeed") in stale
    assert ("news", "HealthyFeed") not in stale


def test_pipeline_writes_injected_report_path_not_production(tmp_path, monkeypatch):
    # Arrange - regression guard: run report writes must respect settings.run_report_path
    # or injected report_path, never clobbering production exports/run_report.json.
    custom_report = tmp_path / "custom_dir" / "report.json"
    monkeypatch.setattr("src.config.settings.run_report_path", str(custom_report))
    prod_path = Path("exports/run_report.json")
    prod_before = prod_path.read_bytes() if prod_path.exists() else None

    # Act 1 - direct call to _write_run_report without output_dir
    path = cli._write_run_report(
        run_id="test-isolation",
        duration_s=5.0,
        status="completed",
        counts={"startups": 1, "products": 1, "papers": 1, "news": 0, "jobs": 0},
        target_count=1,
        resolution_log_rows=0,
        phase1=True,
        phase2=False,
    )

    # Assert 1 - report landed on injected path, production untouched
    assert path == str(custom_report)
    assert custom_report.exists()
    assert json.loads(custom_report.read_text())["run_id"] == "test-isolation"
    prod_after = prod_path.read_bytes() if prod_path.exists() else None
    assert prod_after == prod_before


def test_run_report_duration_completed_and_interrupted(tmp_path):
    # Arrange - test both completed (e.g. ~22 min / 1320.4s) and interrupted runs
    report_completed = tmp_path / "completed_report.json"
    report_interrupted = tmp_path / "interrupted_report.json"

    # Act 1 - completed run duration accounting (~22 min)
    cli._write_run_report(
        run_id="run-22min",
        duration_s=1320.4,
        status="completed",
        counts={"startups": 1000, "products": 1000, "papers": 1000, "news": 39, "jobs": 11},
        target_count=1000,
        resolution_log_rows=1589,
        phase1=True,
        phase2=True,
        report_path=str(report_completed),
    )

    # Act 2 - interrupted run duration accounting (e.g. elapsed 45.3s)
    cli._write_run_report(
        run_id="run-interrupted",
        duration_s=45.3,
        status="interrupted",
        counts={"startups": 250, "products": 180, "papers": 300, "news": 0, "jobs": 0},
        target_count=1000,
        resolution_log_rows=40,
        phase1=True,
        phase2=True,
        report_path=str(report_interrupted),
    )

    # Assert
    data_completed = json.loads(report_completed.read_text())
    assert data_completed["duration_seconds"] == 1320.4
    assert data_completed["status"] == "completed"

    data_interrupted = json.loads(report_interrupted.read_text())
    assert data_interrupted["duration_seconds"] == 45.3
    assert data_interrupted["status"] == "interrupted"


def test_render_summary_surfaces_runtime(capsys):
    # Arrange
    metrics = {"total_nodes": 4567, "total_edges": 1656}
    duration_22min = 1320.0  # 22 minutes

    # Act 1 - render summary with ~22 min duration
    cli._render_summary(
        run_phase1=True,
        run_phase2=True,
        target_count=1000,
        startups=[1] * 1000,
        products=[1] * 1000,
        papers=[1] * 1000,
        news=[1] * 39,
        jobs=[1] * 11,
        metrics=metrics,
        output_xlsx="exports/test.xlsx",
        duration_s=duration_22min,
    )
    captured = capsys.readouterr().out

    # Assert 1 - Total runtime row is surfaced with human-friendly string
    assert "Total runtime" in captured
    assert "~22 min" in captured

    # Act 2 - seconds formatting when < 60s
    cli._render_summary(
        run_phase1=False,
        run_phase2=False,
        target_count=0,
        startups=[],
        products=[],
        papers=[],
        news=[],
        jobs=[],
        metrics=metrics,
        output_xlsx="exports/test.xlsx",
        duration_s=45.2,
    )
    captured_seconds = capsys.readouterr().out
    assert "Total runtime" in captured_seconds
    assert "45.2s" in captured_seconds
