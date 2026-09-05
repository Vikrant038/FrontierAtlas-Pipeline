"""
Unit tests for Excel and Knowledge Graph exporters.
Follows AAA pattern per CODING_STANDARDS.md Pillar 7.
"""

import csv
import os
import tempfile
import openpyxl

from datetime import datetime, timezone

from src.exporters.base import ENTITY_SPECS
from src.exporters.excel_exporter import ExcelExporter
from src.exporters.graph_builder import KnowledgeGraphBuilder
from src.schemas.entities import (
    EntityResolutionLog,
    JobContent,
    JobRecord,
    MatchMethodEnum,
    NewsContent,
    NewsRecord,
    PricingModelEnum,
    ProductContent,
    ProductRecord,
    ResearchPaperContent,
    ResearchPaperRecord,
    RoleFamilyEnum,
    SourceMetadata,
    StartupContent,
    StartupRecord,
)


def test_excel_exporter_creates_6_tabs():
    # Arrange
    exporter = ExcelExporter()
    with tempfile.TemporaryDirectory() as tmp_dir:
        output_file = os.path.join(tmp_dir, "test_output.xlsx")
        
        startups = [
            StartupRecord(
                source=SourceMetadata(name="YC", url="https://yc.com"),
                content=StartupContent(entityName="Anthropic")
            )
        ]
        products = [
            ProductRecord(
                source=SourceMetadata(name="FP", url="https://fp.io"),
                content=ProductContent(
                    startupName="Anthropic",
                    productName="Claude",
                    pricingModel=PricingModelEnum.FREEMIUM
                )
            )
        ]

        # Act
        exporter.export(
            filepath=output_file,
            startups=startups,
            products=products,
        )

        # Assert
        assert os.path.exists(output_file)
        wb = openpyxl.load_workbook(output_file)
        expected_sheets = [
            "Startups", "Products", "Research_Papers", "Jobs_24h", "News_24h", "Entity Mapping Log"
        ]
        assert wb.sheetnames == expected_sheets


def test_excel_exporter_replaces_stale_tab_on_rerun_with_empty_dataset():
    # Arrange: first run writes a populated News_24h tab; second run exports with
    # news=[] (e.g. --phase 1). The stale rows must NOT survive the re-export.
    exporter = ExcelExporter()
    with tempfile.TemporaryDirectory() as tmp_dir:
        output_file = os.path.join(tmp_dir, "test_output.xlsx")
        news = [
            NewsRecord(
                source=SourceMetadata(name="TechCrunch AI", url="https://techcrunch.com/a"),
                content=NewsContent(
                    title="Anthropic announces Claude enterprise features",
                    published_date=datetime.now(timezone.utc),
                    full_text="Some sufficiently long article body text for the schema.",
                ),
            )
        ]
        exporter.export(filepath=output_file, news=news)
        wb = openpyxl.load_workbook(output_file)
        assert wb["News_24h"].max_row == 2

        # Act: re-export same file, no news this run
        exporter.export(filepath=output_file, news=[])

        # Assert: News_24h exists but contains only the header row
        wb2 = openpyxl.load_workbook(output_file)
        assert "News_24h" in wb2.sheetnames
        assert wb2["News_24h"].max_row == 1
        assert wb2["News_24h"].cell(row=2, column=1).value is None


def test_knowledge_graph_builder_links_startup_to_product():
    # Arrange
    builder = KnowledgeGraphBuilder()
    startups = [
        StartupRecord(
            source=SourceMetadata(name="YC", url="https://yc.com"),
            content=StartupContent(entityName="Anthropic")
        )
    ]
    products = [
        ProductRecord(
            source=SourceMetadata(name="FP", url="https://fp.io"),
            content=ProductContent(
                startupName="Anthropic",
                productName="Claude",
                pricingModel=PricingModelEnum.FREEMIUM
            )
        )
    ]

    # Act
    graph = builder.build_graph(startups=startups, products=products)
    metrics = builder.get_summary_metrics()

    # Assert
    assert graph.has_node("Startup:Anthropic")
    assert graph.has_node("Product:Claude")
    assert graph.has_edge("Startup:Anthropic", "Product:Claude")
    assert metrics["total_nodes"] >= 2
    assert metrics["total_edges"] >= 1


def _paper_record(title: str, paper_url: str, github_url=None):
    return ResearchPaperRecord(
        content=ResearchPaperContent(
            title=title,
            authors=["Author One"],
            paper_url=paper_url,
            github_url=github_url,
            github_stars=42 if github_url else None,
            published_date=datetime.now(timezone.utc),
        )
    )


def test_graph_add_papers_builds_implied_repo_edges():
    # Arrange
    builder = KnowledgeGraphBuilder()

    # Act
    builder._add_papers([
        _paper_record("With Repo", "https://arxiv.org/abs/1", github_url="https://github.com/acme/repo"),
        _paper_record("Without Repo", "https://arxiv.org/abs/2"),
    ])

    # Assert: repo link creates Repo node + IMPLEMENTED_IN edge; absent link does not
    assert builder.graph.has_node("Paper:https://arxiv.org/abs/1")
    assert builder.graph.has_node("Repo:https://github.com/acme/repo")
    assert builder.graph.has_edge("Paper:https://arxiv.org/abs/1", "Repo:https://github.com/acme/repo")
    assert builder.graph.has_node("Paper:https://arxiv.org/abs/2")
    assert not any("Repo:" in n for n in builder.graph.nodes if "abs/2" in n)


def test_graph_add_jobs_builds_hired_via_edges():
    # Arrange
    builder = KnowledgeGraphBuilder()
    jobs = [
        JobRecord(
            source=SourceMetadata(name="RemoteOK", url="https://remoteok.com/jobs/1"),
            content=JobContent(
                company="Anthropic",
                title="AI Engineer",
                date=datetime.now(timezone.utc),
                is_remote=True,
                role_family=RoleFamilyEnum.ENGINEERING,
            ),
        )
    ]

    # Act: section builder on its own must create the startup node and the edge
    builder._add_jobs(jobs)

    # Assert
    assert builder.graph.has_node("Startup:Anthropic")
    assert builder.graph.has_edge("Startup:Anthropic", "Job:Anthropic - AI Engineer")


def test_graph_add_news_links_mentioning_startup_or_publisher():
    # Arrange: one headline mentioning an existing startup, one unmatched
    builder = KnowledgeGraphBuilder()
    startups = [StartupRecord(source=SourceMetadata(name="YC", url="https://yc.com"), content=StartupContent(entityName="Acme Corp"))]
    news = [
        NewsRecord(
            source=SourceMetadata(name="TechCrunch", url="https://tc.com/1"),
            content=NewsContent(
                title="Acme Corp ships its agent platform",
                published_date=datetime.now(timezone.utc),
                summary=None,
                full_text="A sufficiently long article body about the launch event today.",
            ),
        ),
        NewsRecord(
            source=SourceMetadata(name="TechCrunch", url="https://tc.com/2"),
            content=NewsContent(
                title="Quantum computing hits a new milestone",
                published_date=datetime.now(timezone.utc),
                summary=None,
                full_text="A sufficiently long article body about the latest research results.",
            ),
        ),
    ]

    # Act
    builder._add_news(news, startups)

    # Assert: matched headline links Startup -> MENTIONED_IN; unmatched falls to Publisher -> PUBLISHED
    assert builder.graph.has_edge("Startup:Acme Corp", "News:Acme Corp ships its agent platform")
    assert builder.graph.has_node("Publisher:TechCrunch")
    assert builder.graph.has_edge("Publisher:TechCrunch", "News:Quantum computing hits a new milestone")


def test_csv_exporter_exports_files():
    # Arrange
    from src.exporters.csv_exporter import CSVExporter
    with tempfile.TemporaryDirectory() as tmp_dir:
        exporter = CSVExporter(output_dir=tmp_dir)
        startups = [
            StartupRecord(
                source=SourceMetadata(name="YC", url="https://yc.com"),
                content=StartupContent(entityName="Anthropic")
            )
        ]

        # Act
        exported = exporter.export_all(startups=startups)

        # Assert
        assert "startups" in exported
        assert os.path.exists(exported["startups"])
        with open(exported["startups"], "r", encoding="utf-8") as f:
            content = f.read()
            assert "Anthropic" in content
            assert "schemaVersion" in content


def test_exports_produce_parseable_files_round_trip(tmp_path):
    # Arrange - Construct complete entity suite across all 6 schemas
    startups = [
        StartupRecord(
            source=SourceMetadata(name="Y Combinator", url="https://ycombinator.com/companies/anthropic"),
            content=StartupContent(
                entityName="Anthropic",
                country="USA",
                batch="W21",
                tags=["AI", "LLM"],
                status="Active",
            ),
        )
    ]
    products = [
        ProductRecord(
            source=SourceMetadata(name="Product Hunt", url="https://producthunt.com/posts/claude-3-5"),
            content=ProductContent(
                startupName="Anthropic",
                productName="Claude 3.5 Sonnet",
                pricingModel=PricingModelEnum.FREEMIUM,
            ),
        )
    ]
    papers = [
        ResearchPaperRecord(
            content=ResearchPaperContent(
                title="Scaling Laws for Neural Language Models",
                authors=["Jared Kaplan", "Sam McCandlish"],
                paper_url="https://arxiv.org/abs/2001.08361",
                github_url="https://github.com/openai/scaling-laws",
                github_stars=1250,
                published_date=datetime(2026, 9, 4, 10, 0, 0, tzinfo=timezone.utc),
            ),
        )
    ]
    jobs = [
        JobRecord(
            source=SourceMetadata(name="RemoteOK AI", url="https://remoteok.com/jobs/101"),
            content=JobContent(
                company="Anthropic",
                title="Staff AI Research Scientist",
                date=datetime(2026, 9, 4, 11, 0, 0, tzinfo=timezone.utc),
                is_remote=True,
                role_family=RoleFamilyEnum.RESEARCH,
            ),
        )
    ]
    news = [
        NewsRecord(
            source=SourceMetadata(name="TechCrunch AI", url="https://techcrunch.com/2026/09/04/anthropic"),
            content=NewsContent(
                title="Anthropic announces Claude enterprise features",
                published_date=datetime(2026, 9, 4, 10, 30, 0, tzinfo=timezone.utc),
                summary="Anthropic announced enterprise controls today.",
                full_text="Anthropic announced new enterprise governance and security features for Claude today.",
            ),
        )
    ]
    logs = [
        EntityResolutionLog(
            rawName="Anthropic, PBC",
            canonicalName="Anthropic",
            matchMethod=MatchMethodEnum.NORMALIZATION_EXACT,
            confidenceScore=0.98,
            sourceUrl="https://ycombinator.com/companies/anthropic",
        )
    ]

    xlsx_path = str(tmp_path / "roundtrip.xlsx")
    csv_dir = str(tmp_path / "csv_out")

    # Act 1 - Export to Excel
    excel_exporter = ExcelExporter()
    excel_exporter.export(
        filepath=xlsx_path,
        startups=startups,
        products=products,
        papers=papers,
        jobs=jobs,
        news=news,
        logs=logs,
    )

    # Act 2 - Export to CSV
    from src.exporters.csv_exporter import CSVExporter
    csv_exporter = CSVExporter(output_dir=csv_dir)
    exported_csvs = csv_exporter.export_all(
        startups=startups,
        products=products,
        papers=papers,
        jobs=jobs,
        news=news,
        logs=logs,
    )

    # Assert 1 - Round-trip read back from Excel with openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    for _, (tab_title, _, expected_headers) in ENTITY_SPECS.items():
        assert tab_title in wb.sheetnames, f"Missing sheet {tab_title}"
        sheet = wb[tab_title]
        actual_headers = [cell.value for cell in sheet[1]]
        assert actual_headers == expected_headers, f"Header mismatch in {tab_title}"
        assert sheet.max_row == 2, f"Expected 1 header + 1 data row in {tab_title}"

    # Spot check specific values in Excel (1-indexed columns)
    assert wb["Startups"].cell(row=2, column=5).value == "Anthropic"
    assert wb["Products"].cell(row=2, column=5).value == "Claude 3.5 Sonnet"
    assert wb["Products"].cell(row=2, column=7).value == "FREEMIUM"
    assert wb["Research_Papers"].cell(row=2, column=3).value == "Scaling Laws for Neural Language Models"
    assert wb["Research_Papers"].cell(row=2, column=7).value == 1250
    assert wb["Jobs_24h"].cell(row=2, column=6).value == "Staff AI Research Scientist"
    assert wb["Jobs_24h"].cell(row=2, column=8).value in (True, "True", "TRUE")
    assert wb["News_24h"].cell(row=2, column=5).value == "Anthropic announces Claude enterprise features"
    assert wb["Entity Mapping Log"].cell(row=2, column=1).value == "Anthropic, PBC"
    assert abs(float(wb["Entity Mapping Log"].cell(row=2, column=5).value) - 0.98) < 1e-4

    # Assert 2 - Round-trip read back from CSV files
    for key, (_, filename, expected_headers) in ENTITY_SPECS.items():
        assert key in exported_csvs
        csv_file = exported_csvs[key]
        assert os.path.exists(csv_file)
        with open(csv_file, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)
        assert len(rows) == 2, f"CSV {filename} must have header + 1 data row"
        assert rows[0] == expected_headers, f"CSV {filename} header mismatch"

