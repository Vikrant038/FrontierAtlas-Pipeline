import os
from typing import List, Optional
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.exporters.base import ENTITY_SPECS
from src.schemas.entities import (
    EntityResolutionLog,
    JobRecord,
    NewsRecord,
    ProductRecord,
    ResearchPaperRecord,
    StartupRecord,
)
from src.utils.logger import logger


class ExcelExporter:
    """Exports pipeline entities into a styled 6-tab Excel workbook using ENTITY_SPECS."""

    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")

    def _add_sheet(self, wb: openpyxl.Workbook, title: str, headers: List[str], records: Optional[List]) -> None:
        """Helper to create, populate, style, and autofit a worksheet."""
        ws = wb.create_sheet(title=title)
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font, cell.fill, cell.alignment = self.HEADER_FONT, self.HEADER_FILL, Alignment(horizontal="center", vertical="center")

        for r in (records or []):
            ws.append(r.to_row() if hasattr(r, "to_row") else r)

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 3, 12), 50)

    def export(
        self,
        filepath: str,
        startups: Optional[List[StartupRecord]] = None,
        products: Optional[List[ProductRecord]] = None,
        papers: Optional[List[ResearchPaperRecord]] = None,
        jobs: Optional[List[JobRecord]] = None,
        news: Optional[List[NewsRecord]] = None,
        logs: Optional[List[EntityResolutionLog]] = None,
    ) -> str:
        """Generate the complete 6-tab workbook."""
        if os.path.exists(filepath):
            try:
                wb = openpyxl.load_workbook(filepath)
            except Exception:
                wb = openpyxl.Workbook()
                wb.remove(wb.active)
        else:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

        datasets = {
            "startups": startups,
            "products": products,
            "papers": papers,
            "jobs": jobs,
            "news": news,
            "logs": logs,
        }

        for key, (title, _, headers) in ENTITY_SPECS.items():
            records = datasets.get(key)
            if records is not None and len(records) > 0:
                if title in wb.sheetnames:
                    wb.remove(wb[title])
                self._add_sheet(wb, title, headers, records)
            elif title not in wb.sheetnames:
                self._add_sheet(wb, title, headers, records or [])

        # Atomic save: write to a temp file and replace so a crash mid-save never
        # corrupts the deliverable workbook.
        tmp_path = f"{filepath}.tmp"
        wb.save(tmp_path)
        os.replace(tmp_path, filepath)
        logger.info(f"Generated 6-tab Excel workbook at {filepath}")
        return filepath
